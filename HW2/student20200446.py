#!/usr/bin/python3

import openpyxl
import math

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

total_list = []
row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value * 1
		ws.cell(row = row_id, column = 7).value = sum_v
		total_list.append([row_id, sum_v])
	row_id += 1

total_list.sort(key=lambda x:x[1], reverse=True)

cnt = len(total_list)

AZ = int(cnt * 0.3)
AP = int(AZ * 0.5)
A0 = AZ - AP

BZ = int(cnt * 0.7)
BP = int((BZ - AZ) * 0.5)
B0 = BZ - AZ - BP

CZ = int(cnt * 1)
CP = int((CZ - BZ) * 0.5)
C0 = CZ - BZ - CP

for i in range(int(CZ)):
	ws.cell(row = total_list[i][0], column = 8).value = 'C0'
for i in range(int(CZ - C0)):
	ws.cell(row = total_list[i][0], column = 8).value = 'C+'
for i in range(int(BZ)):
	ws.cell(row = total_list[i][0], column = 8).value = 'B0'
for i in range(int(BZ - B0)):
	ws.cell(row = total_list[i][0], column = 8).value = 'B+'
for i in range(int(AZ)):
	ws.cell(row = total_list[i][0], column = 8).value = 'A0'
for i in range(int(AP)):
	ws.cell(row = total_list[i][0], column = 8).value = 'A+'
wb.save("student.xlsx")
